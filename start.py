from smart import *
import os
from datetime import datetime
import logging
import openpyxl
from smart.type import AccountType
logger = logging.getLogger()
CONFIG_FILENAME = "grid_target.xlsx" # "I:\smartx_python\gridtrading\grid_target.xlsx"
sz_exchange = smart.Type.Exchange.SZE
sh_exchange = smart.Type.Exchange.SSE

stock_dict = {}

class Stock:
    def __init__(self,strStockCode,strExchange,fInitBasisPrice,fSellPriceDelta,
                    fBuyPriceDelta,fPriceUpperBound,fPriceLowerBound,iAmountPerEntrust,
                    iMaxBuyAmount,iMaxSellAmount,iMaxNettingAmount,index = -1):
        self.strStockCode = strStockCode         # 股票代码 例 000666
        self.strExchange = strExchange           # 交易所 例 SSE
        self.fInitBasisPrice = fInitBasisPrice   # 初始基准价格
        self.fSellPriceDelta = fSellPriceDelta   # 卖出价差，为百分比，如，0.02代表价差百分比为2%
        self.fBuyPriceDelta = fBuyPriceDelta     # 买入价差，为百分比
        self.fPriceUpperBound = fPriceUpperBound # 价格上限
        self.fPriceLowerBound = fPriceLowerBound # 价格下限
        self.iAmountPerEntrust = iAmountPerEntrust # 单次委托数量 todo amount 改为volume
        self.iMaxBuyAmount = iMaxBuyAmount       # 最大买入数量
        self.iMaxSellAmount = iMaxSellAmount     # 最大卖出数量
        self.iMaxNettingAmount = iMaxNettingAmount     # 最大轧差
        self.iBuyAmount = 0                      # 委托买入数量
        self.iSellAmount = 0                     # 委托卖出数量
        self.fCurrBasisPrice = fInitBasisPrice   # 储存上一次的买卖成交均价
        self.isSell = True
        self.isBuy = True
        self.index = index #用于记录在excel中的索引序号,快速保存该股票的最新交易价格

def read_excel(config_file_dir):
    config_file_path =os.path.join( config_file_dir, CONFIG_FILENAME)
    wb = openpyxl.load_workbook(config_file_path) # 读取xlsx文件
    sheet1 = wb.active
    smart.cache.set("account", str(sheet1.cell(1,2).value) )

    sz = []
    sh = []
    nrows = sheet1.max_row + 1 
    ncols = sheet1.max_column + 1
    for i in range(3,nrows):  #
        instrument_id = sheet1.cell(i,1).value
        exchange_id = sheet1.cell(i,2).value
        if isinstance(instrument_id,str) and exchange_id == sz_exchange:
            sz.append(instrument_id)
        elif isinstance(instrument_id,str) and exchange_id == sh_exchange:
            sh.append(instrument_id)
        else:
            logger.warning("warning: error instrument_id or exchange_id info in the %s row." , i)
            continue

        if isinstance(instrument_id,str) and (exchange_id == sz_exchange or exchange_id == sh_exchange): #modified by shizhao on 20191125
            key = instrument_id + exchange_id
            stock_dict[key] = Stock(sheet1.cell(i,1).value, sheet1.cell(i,2).value, float(sheet1.cell(i,3).value), float(sheet1.cell(i,4).value)/100.0, float(sheet1.cell(i,5).value)/100.0, float(sheet1.cell(i,6).value), float(sheet1.cell(i,7).value), sheet1.cell(i,8).value, sheet1.cell(i,9).value, sheet1.cell(i,10).value ,sheet1.cell(i,11).value ,i)
    
    #smart.cache.set("stock_dict", stock_dict )
    smart.cache.set("sz", sz)
    smart.cache.set("sh", sh)


def init():
    try:
        strategy_platform_type = smart.Type.StrategyPlatformType.FrontPy
        strategy_id = None
        price_type = smart.Type.PriceType.Limit
        #smart.Type.Side.Buy
        offset = smart.Type.Offset.Init
        order_client_id = 0
        parent_order_id = ""
        business_type = smart.Type.BusinessType.CASH

        #logger.debug("file_abspath:%s ", os.path.abspath(__file__) )
        #logger.debug("file_dirname:%s ", os.path.dirname(__file__) )


        read_excel(os.path.dirname(__file__))
        account_id = smart.cache["account"]
        #logger.debug("smart.cache.sh: %s",smart.utils.toString(smart.cache["sh"]))
        #logger.debug("smart.cache.sz: %s",smart.utils.toString(smart.cache["sz"]))
        #logger.debug("smart.cache.stock_dict: %s",smart.utils.toString(stock_dict))dddfff
        for key,value in stock_dict.items():
            logger.debug("stock_dict:  key:%s , fInitBasisPrice:%s", key , value.fInitBasisPrice)
        
        # for item in smart.cache["sh"]:
        #     logger.debug("sz: item:%s", item )

        smart.cache.set("begin_time", '093000')   
        smart.cache.set("end_time", '150000')

        smart.subscribe(account_id, smart.cache["sh"] , sh_exchange,False)
        smart.subscribe(account_id, smart.cache["sz"] , sz_exchange,False)
    except Exception as e:
        logger.debug("err: %s", e)

    def on_order(order):
        logger.debug("get on_order: %s",smart.utils.toString(order))
        key = order.instrument_id + order.exchange_id
        if key in stock_dict:
            stock = stock_dict[key]
            if order.status == smart.Type.OrderStatus.Cancelled or order.status ==  smart.Type.OrderStatus.Error or order.status ==  smart.Type.OrderStatus.Filled or order.status ==  smart.Type.OrderStatus.PartialFilledNotActive:
                if order.side ==  smart.Type.Side.Buy:
                    stock.iBuyAmount = stock.iBuyAmount - order.volume_left
                elif order.side ==  smart.Type.Side.Sell:
                    stock.iSellAmount = stock.iSellAmount - order.volume_left
                else:
                    pass
                        
                if order.volume_traded > 0:#有成交股数
                    wb = openpyxl.load_workbook(CONFIG_FILENAME) # 读取xlsx文件
                    ws = wb.active
                    stock.fCurrBasisPrice = order.amount_traded/order.volume_traded #修改最新成交价
                    ws.cell(stock.index,3).value = stock.fCurrBasisPrice
                    wb.save(CONFIG_FILENAME)
                    
                    stock.isSell = True                    
                    stock.isBuy = True
                else:#无成交股数
                    if order.side ==  smart.Type.Side.Buy:#需要修改下
                        stock.isBuy = True
                    else:#取消卖出
                        stock.isSell = True

    smart.current_account.on_order(on_order)

    def on_trade(trade):
        logger.debug("get on_trade: %s",smart.utils.toString(trade))
    smart.current_account.on_trade(on_trade)

    def on_assets(assets):
        logger.debug("get on_assets: %s",smart.utils.toString(assets))
    smart.current_account.on_assets(on_assets)

    def on_position(position):
        logger.debug("get on_position: %s",smart.utils.toString(position))
    smart.current_account.on_position(on_position)

    #撤单
    def cancel_order_callback(data,err):
        logger.debug("get cancel_insert:%s",data)

    #下委托单
    def insert_order_callback(order,err):
        logger.debug("get insert_order: %s",smart.utils.toString(order))

        if order.order_id != '0':###已经考虑到报单错误或者交易所拒单的情况
            key = order.instrument_id + order.exchange_id
            stock = stock_dict[key]
            if order.side == smart.Type.Side.Sell:
                key = order.instrument_id + order.exchange_id
                stock = stock_dict[key]
                stock.iSellAmount = stock.iSellAmount + order.volume
                stock.isSell = False
            elif order.side == smart.Type.Side.Buy:
                stock.iBuyAmount = stock.iBuyAmount + order.volume
                stock.isBuy = False
        #smart.cancel_order(account_id, order_id, cancel_order_callback)
    #smart.insert_order(account_id,strategy_platform_type,strategy_id,instrument_id,exchange_id,price,volume,price_type,side,offset,order_client_id,parent_order_id,business_type,insert_order_callback)
    


    def on_quote(quote):
        logger.debug("quote_time_type:%s",type(quote.data_time))  #str  20230209151145000
            #added by shizhao on 20191125
        curr_time = quote.data_time[8:14]
        if curr_time <= smart.cache["begin_time"] or curr_time >= smart.cache["end_time"]:
            logger.debug("warning:当前时间是%s,不在交易时间区间内！！！",curr_time)
            return
        
        key = quote.instrument_id + quote.exchange_id
        #context.log.info("instrument_id:{} last_price:{}".format(quote.instrument_id, quote.last_price))
        if key in stock_dict:
            stock = stock_dict[key]
            logger.debug("quote: %s",smart.utils.toString(quote))
            if quote.last_price > stock.fCurrBasisPrice:#卖的可能
                rate_of_price_increase = quote.last_price/stock.fCurrBasisPrice - 1.0
                multiple = int(rate_of_price_increase/stock.fSellPriceDelta)
                if 0 == multiple:
                    return
                new_entrust_price = round(stock.fCurrBasisPrice*(1.0 + stock.fSellPriceDelta*multiple), 2)
                if new_entrust_price <= stock.fPriceUpperBound and new_entrust_price <= quote.upper_limit_price and new_entrust_price >= quote.lower_limit_price and stock.isSell and (stock.iSellAmount + stock.iAmountPerEntrust <= stock.iMaxSellAmount) and stock.iSellAmount+stock.iAmountPerEntrust-stock.iBuyAmount <= stock.iMaxNettingAmount:
                    smart.insert_order(account_id,strategy_platform_type,strategy_id,quote.instrument_id, quote.exchange_id , new_entrust_price ,int(stock.iAmountPerEntrust)*multiple,price_type , smart.Type.Side.Sell ,offset,order_client_id,parent_order_id,business_type,insert_order_callback)
            
            elif quote.last_price < stock.fCurrBasisPrice:#买的可能
                rate_of_price_decrease = 1.0 - quote.last_price/stock.fCurrBasisPrice
                multiple = int(rate_of_price_decrease/stock.fBuyPriceDelta)
                if 0 == multiple:
                    return            
                new_entrust_price = round(stock.fCurrBasisPrice*(1.0 - stock.fBuyPriceDelta*multiple), 2)
                if new_entrust_price >= stock.fPriceLowerBound and new_entrust_price >= quote.lower_limit_price and new_entrust_price <= quote.upper_limit_price and stock.isBuy and (stock.iBuyAmount + stock.iAmountPerEntrust <= stock.iMaxBuyAmount) and stock.iBuyAmount+stock.iAmountPerEntrust-stock.iSellAmount <= stock.iMaxNettingAmount:
                    smart.insert_order(account_id,strategy_platform_type,strategy_id,quote.instrument_id, quote.exchange_id , new_entrust_price ,int(stock.iAmountPerEntrust)*multiple, price_type , smart.Type.Side.Buy ,offset,order_client_id,parent_order_id,business_type,insert_order_callback)


            else:#no need any operation
                pass
    smart.on(smart.Event.ON_QUOTE,on_quote)
    

def show():
   logger.debug("show")
def hide():
    logger.debug("hide")
def close():
    logger.debug("close")

smart.on_init(init)
smart.on_show(show)
smart.on_hide(hide)
smart.on_close(close)